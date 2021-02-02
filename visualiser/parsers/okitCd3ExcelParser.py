#!/usr/bin/python

"""Provide Module Description
"""

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~#
__author__ = ["Andrew Hopkinson (Oracle Cloud Solutions A-Team)"]
__version__ = "1.0.0.0"
__module__ = "OkitCD3ExcelParser"

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~#


import getopt
import json
from openpyxl import load_workbook
import sys
import uuid
from common.okitCommon import writeJsonFile
from common.okitCommon import jsonToFormattedString
from common.okitLogging import getLogger
from parsers.okitParser import OkitParser

# Configure logging
logger = getLogger()


class OkitCd3ExcelParser(OkitParser):
    END = "<END>"
    ROOT_COMPARTMENT_ID = "cd3-root-comp"
    COMPARTMENT_SHEET = "Compartments"
    INSTANCE_SHEET = "Instances"
    ROUTE_TABLE_SHEET = "RouteRulesinOCI"
    SUBNET_SHEET = "Subnets"
    VCN_SHEET = "VCNs"

    DYNAMIC_ROUTING_GATEWAY = "dynamic_routing_gateways"
    INTERNET_GATEWAY = "internet_gateways"
    LOCAL_PEERING_GATEWAY = "local_peering_gateways"
    NAT_GATEWAY = "nat_gateways"
    SERVICE_GATEWAY = "service_gateways"

    def __init__(self, cd3_excel=None):
        self.cd3_excel = cd3_excel
        super(OkitCd3ExcelParser, self).__init__(None)

    # Parse Excel Sheets

    def parse(self, cd3_excel=None):
        if cd3_excel is not None:
            self.cd3_excel = cd3_excel

        self.okit_json = self.initialiseOkitJson()
        result_json = {"okit_json": {}, "warnings": {}}

        if self.cd3_excel is not None:
            logger.info("Excel File  {0!s:s}".format(cd3_excel))
            wb = load_workbook(cd3_excel, read_only=True)
            logger.info(wb.sheetnames)
            # Set Root Compartment Id
            self.okit_json["compartments"][0]["id"] = self.ROOT_COMPARTMENT_ID
            self.parseCompartments(wb[self.COMPARTMENT_SHEET])
            self.parseVirtualCloudNetworks(wb[self.VCN_SHEET])
            self.parseRouteTables(wb[self.ROUTE_TABLE_SHEET])
            self.parseSubnets(wb[self.SUBNET_SHEET])
            self.parseInstances(wb[self.INSTANCE_SHEET])
            #logger.info(jsonToFormattedString(self.okit_json))
            result_json["okit_json"] = self.okit_json
            wb.close()

        return result_json

    def parseCompartments(self, sheet):
        return

    def parseInstances(self, sheet):
        row = 3
        while not self.endRow(sheet, row):
            # Destination cidr
            col = "E"
            resource = {
                "compartment_id": self.getCompartmentId(sheet["B{0:d}".format(row)].value),
                "availability_domain": sheet["C{0:d}".format(row)].value[-1],
                "id": self.generateId("instance"),
                "display_name": sheet["D{0:d}".format(row)].value,
                "fault_domain": sheet["E{0:d}".format(row)].value,
                "shape": sheet["J{0:d}".format(row)].value,
                "vnics": [],
                "source_details": {}
            }
            # VNICs
            ## Subnet
            col = "F"
            vnic = {
                "subnet_id": self.getSubnetId(sheet["{0:s}{1:d}".format(col, row)].value),
                "hostname_labal": resource["display_name"].lower()
            }
            ## Public IP
            col = "G"
            #vnic["assign_public_ip"] = (sheet["{0:s}{1:d}".format(col, row)].value.lower() == 'true')
            vnic["assign_public_ip"] = sheet["{0:s}{1:d}".format(col, row)].value
            ## NSGs
            resource["vnics"].append(vnic)
            # Source details
            col = "I"
            source_details = sheet["{0:s}{1:d}".format(col, row)].value.split("::")
            resource["source_details"]["source_type"] = source_details[0]
            resource["source_details"]["os"] = source_details[-1]
            # Add the Instance
            self.okit_json["instances"].append(resource)
            row += 1
        return

    def parseInternetGateways(self, sheet):
        return

    def parseRouteTables(self, sheet):
        row = 3
        while not self.endRow(sheet, row):
            vcn_id = self.getVcnId(sheet["C{0:d}".format(row)].value)
            name = sheet["D{0:d}".format(row)].value
            resource = self.getRouteTable(name, vcn_id)
            if resource is None:
                resource = {
                    "compartment_id": self.getCompartmentId(sheet["B{0:d}".format(row)].value),
                    "vcn_id": vcn_id,
                    "id": self.generateId("routetable"),
                    "display_name": name,
                    "route_rules": []
                }
                self.okit_json["route_tables"].append(resource)
            # Add Rules
            rule = {}
            # Destination cidr
            col = "E"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet["{0:s}{1:d}".format(col, row)].value.strip() != "":
                rule["destination"] = "{0!s:s}".format(str(sheet["{0:s}{1:d}".format(col, row)].value))
            # Destination Name
            col = "F"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None :
                destination = "{0!s:s}".format(str(sheet["{0:s}{1:d}".format(col, row)].value)).strip()
            else:
                destination = ""
            # Target Type
            col = "G"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet["{0:s}{1:d}".format(col, row)].value.strip() != "":
                rule["target_type"] = "{0!s:s}s".format(str(sheet["{0:s}{1:d}".format(col, row)].value).replace(" ", "_")).lower()
                if rule["target_type"] == self.INTERNET_GATEWAY:
                    rule["network_entity_id"] = self.getInternetGatewayId(vcn_id)
                elif rule["target_type"] == self.NAT_GATEWAY:
                    rule["network_entity_id"] = self.getNatGatewayId(vcn_id)
                elif rule["target_type"] == self.SERVICE_GATEWAY:
                    rule["network_entity_id"] = self.getServiceGatewayId(vcn_id)
                elif rule["target_type"] == self.LOCAL_PEERING_GATEWAY:
                    rule["network_entity_id"] = self.getLocalPeeringGatewayId(destination)
                elif rule["target_type"] == self.DYNAMIC_ROUTING_GATEWAY:
                    rule["network_entity_id"] = self.getDynamicRoutingGatewayId(destination)
            resource["route_rules"].append(rule)
            # Increment row
            row += 1
        return

    def parseSubnets(self, sheet):
        row = 3
        while not self.endRow(sheet, row):
            subnet = {
                "compartment_id": self.getCompartmentId(sheet["B{0:d}".format(row)].value),
                "vcn_id": self.getVcnId(sheet["C{0:d}".format(row)].value),
                "id": self.generateId("subnet"),
                "display_name": sheet["D{0:d}".format(row)].value,
                "cidr_block": sheet["E{0:d}".format(row)].value,
                "prohibit_public_ip_on_vnic": sheet["E{0:d}".format(row)].value.strip().lower() == 'private',
                "availability_domain": "0" if sheet["F{0:d}".format(row)].value.strip().lower() == 'regional' else sheet["F{0:d}".format(row)].value.strip()
            }
            # DNS Label
            col = "Q"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet["{0:s}{1:d}".format(col, row)].value.strip() != "":
                subnet["dns_label"] = sheet["{0:s}{1:d}".format(col, row)].value
            # Add the Subnet
            self.okit_json["subnets"].append(subnet)
            row += 1
        return

    def parseVirtualCloudNetworks(self, sheet):
        row = 3
        while sheet["A{0:d}".format(row)].value != self.END:
            vcn = {
                "compartment_id": self.getCompartmentId(sheet["B{0:d}".format(row)].value),
                "id": self.generateId("vcn"),
                "display_name": sheet["C{0:d}".format(row)].value,
                "cidr_block": sheet["D{0:d}".format(row)].value
            }
            # DNS Label
            col = "K"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip() != "":
                vcn["dns_label"] = sheet["{0:s}{1:d}".format(col, row)].value
            # Dynamic Routing Gateway
            col = "E"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip() != '' and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip().lower() != "n":
                logger.info("Create Dynamic Routing Gateway")
                self.createDynamicRoutingGateway(vcn, sheet["{0:s}{1:d}".format(col, row)].value.strip())
            # Internet Gateway
            col = "F"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip() != '' and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip().lower() != "n":
                logger.info("Create Internet Gateway")
                self.createInternetGateway(vcn, sheet["{0:s}{1:d}".format(col, row)].value.strip())
            # NAT Gateway
            col = "G"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip() != '' and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip().lower() != "n":
                logger.info("Create NAT Gateway")
                self.createNatGateway(vcn, sheet["{0:s}{1:d}".format(col, row)].value.strip())
            # Service Gateway
            col = "H"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip() != '' and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip().lower() != "n":
                logger.info("Create Service Gateway")
                self.createServiceGateway(vcn, sheet["{0:s}{1:d}".format(col, row)].value.strip())
            # Local Peering Gateway
            col = "J"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None:
                hub_spoke = sheet["{0:s}{1:d}".format(col, row)].value.strip()
            else:
                hub_spoke = ""
            col = "I"
            if sheet["{0:s}{1:d}".format(col, row)].value is not None and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip() != '' and sheet[
                "{0:s}{1:d}".format(col, row)].value.strip().lower() != "n":
                logger.info("Create Local Peering Gateway")
                self.createLocalPeeringGateway(vcn, sheet["{0:s}{1:d}".format(col, row)].value.strip(), hub_spoke)
            # Add the Virtual Cloud Network
            self.okit_json["virtual_cloud_networks"].append(vcn)
            row += 1
        return

    # Test Functions

    def endRow(self, sheet, row):
        testCells ="ABCDEFGHIJK"
        if sheet["A{0:d}".format(row)].value == self.END:
            return True
        elif all(sheet["{0:s}{1:d}".format(c, row)].value is None or sheet["{0:s}{1:d}".format(c, row)].value.strip() == "" for c in testCells):
            return True
        else:
            return False

    # Create Resources

    def createDynamicRoutingGateway(self, vcn, name):
        if name.lower() == 'y':
            name = "{0!s:s}_drg".format(vcn["display_name"])
        for display_name in name.split(","):
            gw = {
                "compartment_id": vcn["compartment_id"],
                "id": self.generateId("dynamicroutinggateway"),
                "vcn_id": vcn["id"],
                "display_name": display_name
            }
            self.okit_json["dynamic_routing_gateways"].append(gw)
        return

    def createInternetGateway(self, vcn, name):
        ig = {
            "compartment_id": vcn["compartment_id"],
            "id": self.generateId("internetgateway"),
            "vcn_id": vcn["id"],
            "display_name": name if name.lower() != 'y' else "{0!s:s}_ig".format(vcn["display_name"])
        }
        self.okit_json["internet_gateways"].append(ig)
        return

    def createLocalPeeringGateway(self, vcn, name, hub_spoke):
        if name.lower() == 'y':
            name = "{0!s:s}_lpg".format(vcn["display_name"])
        for display_name in name.split(","):
            gw = {
                "compartment_id": vcn["compartment_id"],
                "id": self.generateId("localpeeringgateway"),
                "vcn_id": vcn["id"],
                "display_name": display_name,
                "peer_id": ""
            }
            if "spoke" in hub_spoke:
                spoke = hub_spoke.split(":")
                vcn_id = self.getVcnId(spoke[-1])
                logger.info("{0:s} - {1!s:s}".format(hub_spoke, vcn_id))
                for lpg in self.okit_json["local_peering_gateways"]:
                    if lpg["vcn_id"] == vcn_id and lpg["peer_id"] == '':
                        lpg["peer_id"] = gw["id"]
                        gw["peer_id"] = lpg["id"]
                        logger.info('Assigned {0:s}'.format(lpg["display_name"]))
                        break;

            self.okit_json["local_peering_gateways"].append(gw)
        return

    def createNatGateway(self, vcn, name):
        nat = {
            "compartment_id": vcn["compartment_id"],
            "id": self.generateId("natgateway"),
            "vcn_id": vcn["id"],
            "display_name": name if name.lower() != 'y' else "{0!s:s}_ng".format(vcn["display_name"])
        }
        self.okit_json["nat_gateways"].append(nat)
        return

    def createServiceGateway(self, vcn, name):
        gw = {
            "compartment_id": vcn["compartment_id"],
            "id": self.generateId("servicegateway"),
            "vcn_id": vcn["id"],
            "display_name": name if name.lower() != 'y' else "{0!s:s}_sg".format(vcn["display_name"])
        }
        self.okit_json["service_gateways"].append(gw)
        return

    # Get Resource Ids

    def getCompartmentId(self, name):
        if not any(x["display_name"] == name for x in self.okit_json["compartments"]):
            compartment = {
                "compartment_id": self.ROOT_COMPARTMENT_ID,
                "id": self.generateId("comp"),
                "display_name": name,
                "description": "Imported from CD3 {0!s:s}".format(name)
            }
            self.okit_json["compartments"].append(compartment)
        return [x["id"] for x in self.okit_json["compartments"] if x["display_name"] == name][0]

    def getDynamicRoutingGatewayId(self, name):
        return [x["id"] for x in self.okit_json["dynamic_routing_gateways"] if x["display_name"] == name][0]

    def getInternetGatewayId(self, vcn_id):
        return [x["id"] for x in self.okit_json["internet_gateways"] if x["vcn_id"] == vcn_id][0]

    def getLocalPeeringGatewayId(self, name):
        return [x["id"] for x in self.okit_json["local_peering_gateways"] if x["display_name"] == name][0]

    def getNatGatewayId(self, vcn_id):
        return [x["id"] for x in self.okit_json["nat_gateways"] if x["vcn_id"] == vcn_id][0]

    def getRouteTable(self, name, vcn_id):
        resource = [x for x in self.okit_json["route_tables"] if x["display_name"] == name and x["vcn_id"] == vcn_id]
        return resource[0] if len(resource) > 0 else None

    def getServiceGatewayId(self, vcn_id):
        return [x["id"] for x in self.okit_json["service_gateways"] if x["vcn_id"] == vcn_id][0]

    def getSubnetId(self, name):
        return [x["id"] for x in self.okit_json["subnets"] if x["display_name"] == name][0]

    def getVcnId(self, name):
        return [x["id"] for x in self.okit_json["virtual_cloud_networks"] if x["display_name"] == name][0]

    def generateId(self, type):
        return "okit-{0!s:s}-{1!s:s}".format(type, str(uuid.uuid4()))


# Execute workflow
def processWorkflow(args):
    logger.info(args)
    if args["input"] != None:
        cd3_parser = OkitCd3ExcelParser()
        result = cd3_parser.parse(args["input"])
        writeJsonFile(result["okit_json"], args["output"])
    return


# Set default values for Args
def defaultArgs():
    args = {}
    args['input'] = "./cd3.xlsx"
    args['output'] = "./okit.json"
    return args


# Read Module Arguments
def readargs(opts, args):
    moduleargs = defaultArgs()

    # Read Module Command Line Arguments.
    for opt, arg in opts:
        if opt in ("-i", "--input"):
            moduleargs['input'] = arg
        elif opt in ("-o", "--output"):
            moduleargs['output'] = arg

    return moduleargs


def usage():
    logger.info("Usage:")


# Main processing function
def main(argv):
    # Configure Parameters and Options
    options = 'i:o:'
    longOptions = ['input=', 'output=']
    # Get Options & Arguments
    try:
        opts, args = getopt.getopt(argv, options, longOptions)
        # Read Module Arguments
        moduleargs = readargs(opts, args)
        processWorkflow(moduleargs)
    except getopt.GetoptError:
        usage()
    except Exception as e:
        print('Unknown Exception please check log file')
        logger.exception(e)
        sys.exit(1)

    return


# Main function to kick off processing
if __name__ == "__main__":
    main(sys.argv[1:])
